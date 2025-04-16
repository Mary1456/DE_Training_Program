import pandas as pd
import mysql.connector
from openpyxl.styles import Font


def db_creation(files, db_config, database_name):
    try:
        connection = mysql.connector.connect(
            host=db_config['host'],
            user=db_config['user'],
            password=db_config['password']
        )

        cursor = connection.cursor()

        cursor.execute(f"CREATE DATABASE IF NOT EXISTS {database_name}")
        cursor.execute(f"USE {database_name}")

        for file in files:
            df = pd.read_csv(file)
            table_name = file.split('.')[0]

            column_defs = []
            for col in df.columns:
                if col.lower() == "id":
                    column_defs.append("`id` INT PRIMARY KEY AUTO_INCREMENT")
                else:
                    column_defs.append(f"`{col}` TEXT")

            columns_query = ', '.join(column_defs)
            create_query = f"CREATE TABLE IF NOT EXISTS `{table_name}` ({columns_query})"
            cursor.execute(create_query)

            insert_cols = [col for col in df.columns if col.lower() != 'id']
            insert_placeholders = ', '.join(['%s'] * len(insert_cols))
            insert_query = f"INSERT INTO `{table_name}` ({', '.join(insert_cols)}) VALUES ({insert_placeholders})"

            for _, row in df.iterrows():
                values = [row[col] for col in insert_cols]
                cursor.execute(insert_query, tuple(values))

        connection.commit()
        cursor.close()
        connection.close()
    except mysql.connector.Error as err:
        print(f"MYSQL ERROR  {err}")
    except Exception as e:
        print("Exception > ", e)


def generate_report(dfs, output_file="report.xlsx"):
    users_df = dfs['Users']
    orders_df = dfs['Orders']
    products_df = dfs['Products']
    categories_df = dfs['Categories']

    #clean data
    users_df.dropna(inplace=True)
    orders_df.dropna(inplace=True)
    products_df.dropna(inplace=True)
    categories_df.dropna(inplace=True)

    # aggregate
    # total price
    products_df_renamed = products_df.rename(columns={'id': 'product_id'})
    prdct_ordr = orders_df.merge(products_df_renamed[['product_id', 'price']],
                                on='product_id', how='left')
    
    prdct_ordr['total_price'] = prdct_ordr['quantity'] * prdct_ordr['price_y']
    total_spent = prdct_ordr.groupby('user_id')['total_price'].sum().reset_index()

    users_df = users_df.rename(columns={'id': 'user_id'})
    total_spent = total_spent.merge(users_df, on='user_id', how='left')

    # filter
    # quantity more than 8
    qty_orders = orders_df[orders_df['quantity'] > 8]
    qty_user = qty_orders.merge(users_df, on='user_id')
    quantity = qty_user.sort_values(by='user_id')  

    # merge dfs
    users_df = dfs['Users'].rename(columns={'id': 'user_id'})
    orders_df = dfs['Orders']
    products_df = dfs['Products'].rename(columns={'id': 'product_id', 'name': 'product_name'})
    categories_df = dfs['Categories'].rename(columns={'id': 'category_id', 'name': 'category_name'})

    orders_products = orders_df.merge(products_df, on='product_id', how='left')
    orders_products_categories = orders_products.merge(categories_df, on='category_id', how='left')
    full_data = orders_products_categories.merge(users_df, on='user_id', how='left')
    full_data = full_data.sort_values(by='user_id') 
    full_data = full_data[[
        'user_id', 'first_name', 'last_name', 'product_name', 'category_name', 'quantity', 'order_date'
    ]]


    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        pd.DataFrame().to_excel(writer, sheet_name='Report', index=False)

        workbook = writer.book
        sheet = workbook['Report']

        # Header 1
        sheet['A1'] = "Users who have ordered more than 8 quantity"
        sheet['A1'].font = Font(size=14, bold=True)

        start_row_1 = 2
        quantity[['user_id', 'first_name', 'last_name', 'email',
                             'quantity', 'order_date']].to_excel(
            writer, sheet_name='Report', startrow=start_row_1, index=False)

        # Header 2
        start_row_2 = start_row_1 + len(quantity) + 5
        sheet.cell(row=start_row_2, column=1).value = "Total money spent by each user"
        sheet.cell(row=start_row_2, column=1).font = Font(size=14, bold=True)

        total_spent[['user_id', 'first_name', 'total_price']].to_excel(
            writer, sheet_name='Report', startrow=start_row_2 + 1, index=False)
        
        # Header 3
        start_row_3 = start_row_2 + len(total_spent) + 5
        sheet.cell(row=start_row_3, column=1).value = "Merging DF"
        sheet.cell(row=start_row_3, column=1).font = Font(size=14, bold=True)

        full_data[['user_id', 'first_name', 'last_name', 'product_name',
                   'category_name', 'quantity', 'order_date']].to_excel(
            writer, sheet_name='Report', startrow=start_row_3 + 1, index=False)

    print(f"Report generated as '{output_file}'")


